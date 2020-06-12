# from openpyxl import Workbook
# from openpyxl.compat import range
# from openpyxl.utils import get_column_letter
#
# wb = Workbook()
# dest_filename = 'empty_book.xlsx'
# ws1 = wb.active
# ws1.title = "range names"
# for row in range(1, 40):
#     ws1.append(range(600))
#
# ws2 = wb.create_sheet(title="Pi")
#
# ws2['F5'] = 3.14
#
# ws3 = wb.create_sheet(title="Data")
# for row in range(10, 20):
#     for col in range(27, 54):        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
#
#     print(ws3['AA10'].value)
# wb.save(filename=dest_filename)
#


from openpyxl import Workbook

wb = Workbook()

ws = wb.active
ws1 = wb.create_sheet('Folha1')
ws2 = wb.create_sheet('Folha2', 0)
ws3 = wb.create_sheet('Folha30', 1)

std = wb.get_sheet_by_name('Folha2')

wb.remove_sheet(std)

print(wb.active)

ws.title = 'New Title'

print(wb.sheetnames)

wb.active = 1

print(wb.active)
