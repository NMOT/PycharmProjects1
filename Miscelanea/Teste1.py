from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws.title = 'Folha1'

ws.sheet_properties.tabColor = "1072BA"

ws['A1'] = 42
ws.append([1,2,3])

ws.append([10,11,12,234,])

# import datetime
#
# ws['A2'] = datetime.datetime.now()

for sheet in wb:
    print(sheet.title)

ws['A3'] = 255

c = ws['A3']

ws['A2'] = 123
ws.cell(row = 2, column = 1)

n = ws['A10']
print(c.value)
print(ws.cell(row = 2, column = 1).value )

print(ws.columns)

n.value = 10000

wb.save('sample.xlsx')