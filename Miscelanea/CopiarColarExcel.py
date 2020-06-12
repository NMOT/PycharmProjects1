# from openpyxl import Workbook
#
# wb = Workbook()
#
# ws = wb.active
#
# ws['A1'] = 'merda'
#
# wb.save('teste.xlsx')

from openpyxl import load_workbook

wb=load_workbook('ActualizaçãoPreços.xlsx')

# ws=wb.active

ws= wb['Folha1']

ws2 = wb['Folha2']

# print(wb.get_sheet_names())

print(ws['F1'].value)

ws['A1'] = 12

print(ws['A1'].value)

#print(ws2['A1'].value)

print(ws['A3'].number_format)

ws['A1'] = 5
ws['A2'] = 10
# soma = ws['A1'] + ws['A2']
# ws['A3'] = soma
# #
# print(ws['A3'.value])

wb.save('ActualizaçãoPreços.xlsx')
