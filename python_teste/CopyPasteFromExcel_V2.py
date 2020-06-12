import glob
import os
import xlrd
import openpyxl
import winsound

list_of_files = glob.glob("C:\\Nkmix\\ForExcel\\*.xls")
latest_file = max(list_of_files, key=os.path.getctime)

print(latest_file)

Orig_book = xlrd.open_workbook(filename=latest_file)

Orig_sheet = Orig_book.sheet_by_name('Matérias-Primas a utilizar')

Orig_data_col1 = [Orig_sheet.cell_value(r, 0) for r in range(Orig_sheet.nrows) if Orig_sheet.cell_value(r, 0) != ""]
Orig_data_col1.pop(0)
Orig_data_col1.pop(0)

Orig_data_col2 = [Orig_sheet.cell_value(r, 1) for r in range(Orig_sheet.nrows) if Orig_sheet.cell_value(r, 1) != ""]
Orig_data_col2.pop(0)

Orig_data_col3 = [Orig_sheet.cell_value(r, 2) for r in range(Orig_sheet.nrows) if Orig_sheet.cell_value(r, 2) != ""]
Orig_data_col3.pop(0)

DestPath = "Y:\\DadosActualizaçãoPowerBi\\Indicadores_Unidade_Alimentos_Compostos_2019.xlsx"


Dest_book = openpyxl.load_workbook(filename=DestPath, data_only=False)

Dest_sheet = Dest_book.get_sheet_by_name('Cons_Est')

for mycol in range(14,17):
    for myrow in range(6,185):
        cellRef = Dest_sheet.cell(row=myrow,column=mycol)
        cellRef.value = ""
for myrow in range(0, int(len(Orig_data_col1))):
    cellRef=Dest_sheet.cell(row=myrow+6,column=14)
    cellRef.value = Orig_data_col1[myrow]

for myrow in range(0, int(len(Orig_data_col1))):
    cellRef=Dest_sheet.cell(row=myrow+6,column=15)
    cellRef.value = Orig_data_col2[myrow]

for myrow in range(0, int(len(Orig_data_col1))):
    cellRef=Dest_sheet.cell(row=myrow+6,column=16)
    cellRef.value = Orig_data_col3[myrow]

Dest_book.save(DestPath)

winsound.MessageBeep(20)
winsound.MessageBeep(20)